"""
二维码文件处理API端点
支持Excel区域序列化、文件序列化、QR码生成和扫描恢复功能
"""
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, Form, Body, Query
from fastapi.responses import JSONResponse, FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
import os
import base64
import zlib
import pickle
import struct
import qrcode
from PIL import Image, ImageDraw, ImageFont
import io
import tempfile
import uuid
import shutil
import cv2  # 导入OpenCV
from collections import OrderedDict
from datetime import datetime, timedelta
import time
import threading
import asyncio
import numpy as np

from schemas.user import User
from core.auth import delete_qrfile,download_qrfile,get_current_user
from db.session import get_db
from config import settings
from core.logger import get_logger
from models.qrfile import QRFile
from sqlalchemy import select, update, func

logger = get_logger(__name__)

router = APIRouter()

# 创建二维码处理目录
QR_DIR = os.path.join(settings.DATA_DIR, "qrfiles")
os.makedirs(QR_DIR, exist_ok=True)

class VideoQRScanner:
    """视频二维码扫描器"""

    def __init__(self, video_path, output_dir):
        self.video_path = video_path
        self.output_dir = output_dir
        self.cap = None
        self.unique_qrs = OrderedDict()  # 使用有序字典存储唯一二维码
        self.frame_count = 0
        self.scanned_frames = 0
        self.scan_interval = 2  # 减少扫描间隔，提高检测率
        self.debug_dir = os.path.join(output_dir, "debug_frames")
        os.makedirs(self.debug_dir, exist_ok=True)

    def start(self):
        """开始扫描视频"""
        try:
            self.cap = cv2.VideoCapture(self.video_path)
            if not self.cap.isOpened():
                raise ValueError("无法打开视频文件")

            self.frame_count = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
            fps = self.cap.get(cv2.CAP_PROP_FPS)
            width = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            
            logger.info(f"开始扫描视频: {os.path.basename(self.video_path)}")
            logger.info(f"视频信息: 总帧数={self.frame_count}, FPS={fps}, 分辨率={width}x{height}")
            
            return self.process_video()
        except Exception as e:
            logger.error(f"视频扫描错误: {str(e)}")
            raise ValueError(f"视频扫描错误: {str(e)}")

    def process_video(self):
        """处理视频帧"""
        try:
            frame_index = 0
            while self.cap.isOpened():
                ret, frame = self.cap.read()
                if not ret:
                    break
                
                frame_index += 1
                self.scanned_frames += 1
                
                # 每隔一定帧数处理一次，减少计算量但不要太大
                if frame_index % self.scan_interval != 0:
                    continue
                
                # 保存一些帧用于调试
                if frame_index % 30 == 0 or frame_index <= 10:
                    debug_path = os.path.join(self.debug_dir, f"frame_{frame_index}.jpg")
                    cv2.imwrite(debug_path, frame)
                
                # 尝试多种图像处理方式提高识别率
                qr_data_list = []
                
                # 方法1: 直接处理原始帧
                qr_data_list.extend(self.decode_frame(frame, f"原始帧 {frame_index}"))
                
                # 方法2: 调整对比度和亮度
                adjusted = self.adjust_contrast_brightness(frame, 1.5, 10)
                qr_data_list.extend(self.decode_frame(adjusted, f"调整对比度帧 {frame_index}"))
                
                # 方法3: 灰度图像
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                qr_data_list.extend(self.decode_frame(gray, f"灰度帧 {frame_index}"))
                
                # 方法4: 二值化
                _, binary = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY)
                qr_data_list.extend(self.decode_frame(binary, f"二值化帧 {frame_index}"))
                
                # 处理检测到的二维码
                for qr_data in qr_data_list:
                    if qr_data and qr_data not in self.unique_qrs:
                        self.unique_qrs[qr_data] = True
                        logger.info(f"在第 {frame_index} 帧发现新二维码，当前共 {len(self.unique_qrs)} 个")
                
                # 每处理100帧输出一次进度
                if frame_index % 100 == 0:
                    progress = min(100, int((frame_index / self.frame_count) * 100))
                    logger.info(f"视频扫描进度: {progress}%, 已处理 {frame_index}/{self.frame_count} 帧, 发现 {len(self.unique_qrs)} 个二维码")
            
            # 扫描完成
            self.cap.release()
            logger.info(f"视频扫描完成，共处理 {self.scanned_frames} 帧，发现 {len(self.unique_qrs)} 个唯一二维码")
            
            # 如果没有找到二维码，记录更详细的信息
            if not self.unique_qrs:
                logger.warning(f"未在视频中发现二维码，请检查视频质量或二维码是否清晰可见")
                logger.warning(f"调试帧保存在: {self.debug_dir}")
            
            # 返回扫描结果
            return list(self.unique_qrs.keys())
        except Exception as e:
            logger.error(f"处理视频帧失败: {str(e)}")
            if self.cap and self.cap.isOpened():
                self.cap.release()
            raise ValueError(f"处理视频帧失败: {str(e)}")
    
    def decode_frame(self, frame, frame_desc):
        """解码单个帧中的二维码"""
        try:
            from pyzbar import pyzbar
            
            # 确保帧是PIL图像
            if isinstance(frame, np.ndarray):
                if len(frame.shape) == 3 and frame.shape[2] == 3:  # 彩色图像
                    pil_image = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
                else:  # 灰度或二值图像
                    pil_image = Image.fromarray(frame)
            else:
                pil_image = frame
            
            # 解码二维码
            decoded_objects = pyzbar.decode(pil_image)
            
            # 提取二维码数据
            result = []
            for obj in decoded_objects:
                if obj.type == 'QRCODE':
                    qr_data = obj.data.decode('utf-8')
                    result.append(qr_data)
                    logger.info(f"在{frame_desc}中检测到二维码")
            
            return result
        except Exception as e:
            logger.warning(f"解码帧失败: {str(e)}")
            return []
    
    def adjust_contrast_brightness(self, img, contrast=1.0, brightness=0):
        """调整图像对比度和亮度"""
        return cv2.convertScaleAbs(img, alpha=contrast, beta=brightness)

class QRProcessor:
    def __init__(self, output_dir):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def serialize_excel_region(self, excel_data, region, sheet_name=None, version=8, progress_callback=None):
        """序列化Excel区域"""
        try:
            from openpyxl import load_workbook
            
            # 创建临时文件
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                temp_file.write(excel_data)
                excel_path = temp_file.name
            
            # 加载Excel
            wb = load_workbook(excel_path)
            
            # 选择sheet
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active
                sheet_name = ws.title
            
            # 解析区域坐标
            min_col, min_row, max_col, max_row = self.parse_region(region)
            
            # 收集数据
            data = {
                'data': [],
                'styles': [],
                'merged': [m.coord for m in ws.merged_cells.ranges],
                'meta': {
                    'source': "uploaded_file.xlsx",
                    'sheet': sheet_name,
                    'region': region,
                    'version': version,
                    'timestamp': datetime.now().isoformat(),
                    'mode': 'region'
                }
            }
            
            # 处理每行数据
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                row_data = []
                row_styles = []
                
                for cell in row:
                    # 处理不同类型的数据
                    if isinstance(cell.value, str) and len(cell.value) > 1000:
                        row_data.append(cell.value[:1000] + "...[TRUNCATED]")
                    else:
                        row_data.append(cell.value)
                    row_styles.append(self.get_style(cell))
                
                data['data'].append(row_data)
                data['styles'].append(row_styles)
            
            # 序列化并压缩
            serialized = pickle.dumps(data)
            compressed = zlib.compress(serialized)
            
            # 添加校验和
            checksum = zlib.crc32(compressed)
            compressed_with_checksum = struct.pack(">I", checksum) + compressed
            
            # 清理临时文件
            try:
                os.unlink(excel_path)
            except:
                pass
                
            return compressed_with_checksum
            
        except Exception as e:
            logger.error(f"Excel序列化失败: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Excel序列化失败: {str(e)}")

    def serialize_file(self, file_data, version=8):
        """序列化任意文件"""
        try:
            # 压缩数据
            compressed = zlib.compress(file_data)
            
            # 添加校验和
            checksum = zlib.crc32(compressed)
            compressed_with_checksum = struct.pack(">I", checksum) + compressed
            
            # 添加文件模式标记
            file_marker = b"FILE_MODE:"
            final_data = file_marker + compressed_with_checksum
            
            return final_data
            
        except Exception as e:
            logger.error(f"文件序列化失败: {str(e)}")
            raise HTTPException(status_code=500, detail=f"文件序列化失败: {str(e)}")

    def create_qr_codes(self, data, max_size=1800, version=8, mode="file"):
        """生成二维码序列"""
        try:
            # 计算base64编码后的最大原始数据大小
            max_raw_size = int(max_size * 0.7)  # 考虑base64开销
            
            # 如果数据很小，直接生成单个二维码
            if len(data) <= max_raw_size:
                # 使用base64编码
                base64_data = base64.b64encode(data).decode('utf-8')
                img = self.create_single_qr(base64_data, f"{mode}")
                
                # 转换为二进制
                img_byte_arr = io.BytesIO()
                img.save(img_byte_arr, format='PNG')
                img_byte_arr.seek(0)
                
                return [{"name": "single", "data": img_byte_arr.getvalue()}]
            
            # 计算需要多少分块
            total_chunks = (len(data) + max_raw_size - 1) // max_raw_size
            chunks = []
            
            # 大数据分块处理
            for i in range(total_chunks):
                start = i * max_raw_size
                end = min(start + max_raw_size, len(data))
                chunk_data = data[start:end]
                
                # 添加分块头并使用base64编码
                header = f"QR:{i + 1}/{total_chunks}|v{version}|{mode}|"
                base64_chunk = base64.b64encode(chunk_data).decode('utf-8')
                
                # 检查总长度
                full_chunk = header + base64_chunk
                if len(full_chunk) > max_size:
                    # 如果超出，减小分块大小
                    new_max_raw_size = int(max_raw_size * 0.9)
                    return self.create_qr_codes(data, new_max_raw_size, version, mode)
                
                # 创建二维码
                name = f"chunk_{i + 1}_of_{total_chunks}"
                img = self.create_single_qr(full_chunk, f"{i + 1}/{total_chunks}")
                
                # 转换为二进制
                img_byte_arr = io.BytesIO()
                img.save(img_byte_arr, format='PNG')
                img_byte_arr.seek(0)
                
                chunks.append({"name": name, "data": img_byte_arr.getvalue()})
            
            return chunks
            
        except Exception as e:
            logger.error(f"创建二维码失败: {str(e)}")
            raise HTTPException(status_code=500, detail=f"创建二维码失败: {str(e)}")

    def restore(self, data):
        """从数据恢复文件"""
        try:
            # 检查是否为文件模式
            if data.startswith(b"FILE_MODE:"):
                return self.restore_file(data[10:])
            else:
                return self.restore_excel_region(data)
                
        except Exception as e:
            logger.error(f"恢复失败: {str(e)}")
            raise HTTPException(status_code=500, detail=f"恢复失败: {str(e)}")

    def restore_excel_region(self, data):
        """恢复Excel区域数据"""
        from openpyxl import Workbook
        
        # 验证数据完整性
        if len(data) < 4:
            raise ValueError("数据过短，无法恢复")
            
        # 提取校验和
        stored_checksum = struct.unpack(">I", data[:4])[0]
        actual_data = data[4:]
        
        # 验证校验和
        actual_checksum = zlib.crc32(actual_data)
        if stored_checksum != actual_checksum:
            raise ValueError(f"数据校验失败: {stored_checksum} != {actual_checksum}")
            
        # 尝试解压数据
        try:
            decompressed = zlib.decompress(actual_data)
        except zlib.error as e:
            # 尝试不解压直接使用
            try:
                decompressed = actual_data
            except:
                raise ValueError(f"解压失败: {str(e)}")
                
        # 反序列化
        try:
            restored = pickle.loads(decompressed)
        except pickle.UnpicklingError as e:
            raise ValueError(f"反序列化失败: {str(e)}")
            
        # 检查模式
        meta = restored.get('meta', {})
        if meta.get('mode') != 'region':
            raise ValueError("数据模式不匹配，请使用文件模式恢复")
            
        # 检查版本兼容性
        data_version = meta.get('version', 0)
        if data_version < 4:
            raise ValueError(f"不兼容的数据版本: {data_version} (需要4+)")
            
        # 创建新工作簿
        wb = Workbook()
        ws = wb.active
        
        # 设置sheet名称
        sheet_name = meta.get('sheet', 'Restored')
        if sheet_name:
            ws.title = sheet_name[:30]  # Excel sheet名称长度限制
            
        # 恢复数据
        for r, (row_data, row_styles) in enumerate(zip(restored['data'], restored['styles'])):
            for c, (value, style) in enumerate(zip(row_data, row_styles)):
                cell = ws.cell(row=r + 1, column=c + 1, value=value)
                self.apply_style(cell, style)
                
        # 恢复合并单元格
        for merged in restored.get('merged', []):
            try:
                ws.merge_cells(merged)
            except:
                continue
                
        # 生成临时文件路径
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"restored_excel_{timestamp}.xlsx"
        output_path = os.path.join(self.output_dir, filename)
        
        # 保存文件
        wb.save(output_path)
        return output_path, filename

    def restore_file(self, data):
        """恢复任意文件"""
        # 验证数据完整性
        if len(data) < 4:
            raise ValueError("数据过短，无法恢复")
            
        # 提取校验和
        stored_checksum = struct.unpack(">I", data[:4])[0]
        actual_data = data[4:]
        
        # 验证校验和
        actual_checksum = zlib.crc32(actual_data)
        if stored_checksum != actual_checksum:
            raise ValueError(f"数据校验失败: {stored_checksum} != {actual_checksum}")
            
        # 解压数据
        try:
            decompressed = zlib.decompress(actual_data)
        except zlib.error as e:
            raise ValueError(f"解压失败: {str(e)}")
            
        # 生成临时文件路径
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"restored_file_{timestamp}"
        output_path = os.path.join(self.output_dir, filename)
        
        # 保存文件
        with open(output_path, 'wb') as f:
            f.write(decompressed)
            
        return output_path, filename

    def create_single_qr(self, data, counter=None):
        """创建单个二维码"""
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(data)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # 添加标记文本
        if counter:
            draw = ImageDraw.Draw(img)
            try:
                # 尝试使用系统字体
                try:
                    font = ImageFont.truetype("arial.ttf", 16)
                except:
                    try:
                        font = ImageFont.truetype("Arial.ttf", 16)
                    except:
                        font = ImageFont.load_default()
                
                text = counter
                bbox = draw.textbbox((0, 0), text, font=font)
                text_w = bbox[2] - bbox[0]
                text_h = bbox[3] - bbox[1]
                
                # 在右下角添加文本
                draw.rectangle(
                    [(img.width - text_w - 10, img.height - text_h - 10),
                     (img.width, img.height)],
                    fill="white"
                )
                draw.text(
                    (img.width - text_w - 5, img.height - text_h - 5),
                    text,
                    font=font,
                    fill="black"
                )
            except Exception as e:
                logger.warning(f"添加二维码文本失败: {e}")
                
        return img

    def combine_data(self, chunks):
        """合并分块数据"""
        # 提取所有分块数据
        chunks_dict = {}
        total_chunks = 0
        current_chunks = 0
        
        # 首先收集所有分块信息
        for chunk in chunks:
            if chunk.startswith("QR:"):
                # 分块格式: "QR:2/5|v8|mode|base64数据"
                parts = chunk.split('|', 3)
                if len(parts) < 4:
                    continue
                    
                header = parts[0]
                data_part = parts[3]
                
                # 解析分块头: "QR:2/5"
                chunk_info = header.split(':')[1]
                chunk_num, total = chunk_info.split('/')
                
                chunks_dict[int(chunk_num)] = data_part
                total_chunks = int(total)
                current_chunks += 1
            else:
                # 单个二维码情况
                return base64.b64decode(chunk)
                
        # 检查是否收集到所有分块
        if current_chunks != total_chunks:
            missing = [i for i in range(1, total_chunks + 1) if i not in chunks_dict]
            raise ValueError(f"数据不完整: 缺少分块 {missing}")
            
        # 按顺序组合分块
        combined_b64 = ''.join(chunks_dict[i] for i in sorted(chunks_dict.keys()))
        return base64.b64decode(combined_b64)

    # 辅助方法
    def parse_region(self, region):
        """解析区域坐标"""
        import re
        
        # 移除空格并转换为大写
        region = region.replace(" ", "").upper()
        
        if ':' in region:
            start, end = region.split(':', 1)
        else:
            start = end = region
            
        # 使用正则表达式提取列和行
        pattern = r"([A-Z]+)(\d+)"
        start_match = re.match(pattern, start)
        end_match = re.match(pattern, end)
        
        if not start_match or not end_match:
            raise ValueError(f"无效的区域格式: {region}")
            
        start_col = start_match.group(1)
        start_row = int(start_match.group(2))
        end_col = end_match.group(1)
        end_row = int(end_match.group(2))
        
        # 列字母转数字
        def col_to_num(col):
            num = 0
            for c in col:
                if c.isalpha():
                    num = num * 26 + (ord(c) - ord('A')) + 1
            return num
            
        return (
            col_to_num(start_col),
            start_row,
            col_to_num(end_col),
            end_row
        )

    def get_style(self, cell):
        """获取单元格样式"""
        try:
            return {
                'font': self.copy_font(cell.font) if cell.font else None,
                'fill': self.copy_fill(cell.fill) if cell.fill else None,
                'border': self.copy_border(cell.border) if cell.border else None,
                'alignment': self.copy_alignment(cell.alignment) if cell.alignment else None,
                'format': cell.number_format
            }
        except Exception:
            return {}

    def apply_style(self, cell, style):
        """应用样式到单元格"""
        from openpyxl.styles import Font, PatternFill, Border, Alignment
        
        if style.get('font'):
            try:
                cell.font = style['font']
            except Exception:
                pass
        if style.get('fill'):
            try:
                cell.fill = style['fill']
            except Exception:
                pass
        if style.get('border'):
            try:
                cell.border = style['border']
            except Exception:
                pass
        if style.get('alignment'):
            try:
                cell.alignment = style['alignment']
            except Exception:
                pass
        if style.get('format'):
            try:
                cell.number_format = style['format']
            except Exception:
                pass

    def copy_font(self, font):
        from openpyxl.styles import Font
        if not font: return None
        return Font(
            name=font.name, size=font.size, bold=font.bold,
            italic=font.italic, strike=font.strike,
            color=font.color
        )

    def copy_fill(self, fill):
        from openpyxl.styles import PatternFill
        if not fill: return None
        return PatternFill(
            fill_type=fill.fill_type,
            start_color=fill.start_color,
            end_color=fill.end_color
        )

    def copy_border(self, border):
        from openpyxl.styles import Border
        if not border: return None
        return Border(
            left=border.left, right=border.right,
            top=border.top, bottom=border.bottom
        )

    def copy_alignment(self, alignment):
        from openpyxl.styles import Alignment
        if not alignment: return None
        return Alignment(
            horizontal=alignment.horizontal,
            vertical=alignment.vertical,
            wrap_text=alignment.wrap_text,
            shrink_to_fit=alignment.shrink_to_fit,
            indent=alignment.indent
        )


@router.post("/serialize-excel")
async def serialize_excel(
    excel_file: UploadFile = File(...),
    region: str = Form(...),
    sheet_name: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    序列化Excel区域为二维码数据
    """
    try:
        # 读取Excel文件内容
        excel_data = await excel_file.read()
        
        # 创建处理器并序列化
        processor = QRProcessor(QR_DIR)
        serialized_data = processor.serialize_excel_region(excel_data, region, sheet_name)
        
        # 生成唯一标识
        session_id = str(uuid.uuid4())
        session_path = os.path.join(QR_DIR, f"{session_id}.data")
        
        # 保存序列化数据
        with open(session_path, 'wb') as f:
            f.write(serialized_data)
            
        # 保存到数据库
        qr_file = QRFile(
            filename=f"{session_id}.data",
            filepath=session_path,
            filetype="excel",
            filesize=len(serialized_data),
            session_id=session_id,
            mode="region",
            original_filename=excel_file.filename,
            original_filesize=len(excel_data),
            user_id=current_user.id,
            description=f"Excel区域: {region}, Sheet: {sheet_name or '默认'}"
        )
        db.add(qr_file)
        await db.commit()
            
        logger.info(f"用户 {current_user.username} 成功序列化Excel区域, ID: {session_id}")
        
        return {
            "success": True,
            "session_id": session_id,
            "data_size": len(serialized_data),
            "mode": "region"
        }
    except Exception as e:
        logger.error(f"序列化Excel失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"序列化Excel失败: {str(e)}")


@router.post("/serialize-file")
async def serialize_file(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    序列化文件为二维码数据
    """
    try:
        # 检查文件大小
        file_data = await file.read()
        if len(file_data) > 10 * 1024 * 1024:  # 10MB
            raise HTTPException(status_code=400, detail="文件过大，建议使用小于10MB的文件")
            
        # 创建处理器并序列化
        processor = QRProcessor(QR_DIR)
        serialized_data = processor.serialize_file(file_data)
        
        # 生成唯一标识
        session_id = str(uuid.uuid4())
        session_path = os.path.join(QR_DIR, f"{session_id}.data")
        
        # 保存序列化数据
        with open(session_path, 'wb') as f:
            f.write(serialized_data)
            
        # 保存到数据库
        qr_file = QRFile(
            filename=f"{session_id}.data",
            filepath=session_path,
            filetype=file.content_type,
            filesize=len(serialized_data),
            session_id=session_id,
            mode="file",
            original_filename=file.filename,
            original_filesize=len(file_data),
            user_id=current_user.id,
            description=f"文件: {file.filename}"
        )
        db.add(qr_file)
        await db.commit()
            
        logger.info(f"用户 {current_user.username} 成功序列化文件 {file.filename}, ID: {session_id}")
        
        return {
            "success": True,
            "session_id": session_id,
            "data_size": len(serialized_data),
            "mode": "file",
            "filename": file.filename
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"序列化文件失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"序列化文件失败: {str(e)}")


@router.post("/generate-qrcodes")
async def generate_qrcodes(
    session_id: str = Form(...),
    chunk_size: int = Form(1800),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    生成二维码序列
    """
    try:
        # 检查会话数据是否存在
        session_path = os.path.join(QR_DIR, f"{session_id}.data")
        if not os.path.exists(session_path):
            raise HTTPException(status_code=404, detail="会话数据不存在，请先序列化数据")
            
        # 读取序列化数据
        with open(session_path, 'rb') as f:
            serialized_data = f.read()
            
        # 检测模式
        mode = "file" if serialized_data.startswith(b"FILE_MODE:") else "region"
        
        # 创建处理器并生成二维码
        processor = QRProcessor(QR_DIR)
        qr_images = processor.create_qr_codes(serialized_data, max_size=chunk_size, mode=mode)
        
        # 创建二维码存储目录
        qr_dir = os.path.join(QR_DIR, session_id)
        os.makedirs(qr_dir, exist_ok=True)
        
        # 保存二维码图片
        image_paths = []
        for i, qr_data in enumerate(qr_images):
            img_path = os.path.join(qr_dir, f"{qr_data['name']}.png")
            with open(img_path, 'wb') as f:
                f.write(qr_data['data'])
            image_paths.append(img_path)
            
        # 更新数据库中的记录
        stmt = update(QRFile).where(QRFile.session_id == session_id).values(
            chunk_count=len(qr_images)
        )
        await db.execute(stmt)
        await db.commit()
            
        logger.info(f"用户 {current_user.username} 成功生成 {len(qr_images)} 个二维码")
        
        return {
            "success": True,
            "session_id": session_id,
            "qr_count": len(qr_images),
            "qr_images": [{"name": os.path.basename(path), "path": path} for path in image_paths]
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"生成二维码失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"生成二维码失败: {str(e)}")


@router.get("/qrcode/{session_id}/{name}")
async def get_qrcode(
    session_id: str,
    name: str,
    # current_user: User = Depends(get_current_user),
):
    """
    获取生成的二维码图片
    """
    try:
        image_path = os.path.join(QR_DIR, session_id, name)
        print(os.path.abspath(image_path))
        if not os.path.exists(image_path):
            raise HTTPException(status_code=404, detail="二维码图片不存在")
            
        return FileResponse(image_path, media_type="image/png")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取二维码失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"获取二维码失败: {str(e)}")


@router.post("/scan-restore")
async def scan_restore(
    files: List[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
):
    """
    扫描二维码图片恢复数据
    """
    try:
        # 创建临时目录
        temp_dir = tempfile.mkdtemp(dir=QR_DIR)
        
        try:
            # 保存上传的图片
            file_paths = []
            for file in files:
                file_path = os.path.join(temp_dir, file.filename)
                with open(file_path, 'wb') as f:
                    f.write(await file.read())
                file_paths.append(file_path)
                
            # 解码所有二维码
            from pyzbar import pyzbar
            
            chunks = []
            for file_path in file_paths:
                try:
                    img = Image.open(file_path)
                    results = pyzbar.decode(img)
                    for result in results:
                        if result.type == 'QRCODE':
                            chunks.append(result.data.decode('utf-8'))
                except Exception as e:
                    logger.warning(f"解码二维码失败: {file_path}, 错误: {str(e)}")
                    
            if not chunks:
                raise HTTPException(status_code=400, detail="未找到有效的二维码数据")
                
            # 合并数据
            processor = QRProcessor(QR_DIR)
            combined_data = processor.combine_data(chunks)
            
            # 恢复数据
            output_path, filename = processor.restore(combined_data)
            
            logger.info(f"用户 {current_user.username} 成功恢复文件: {filename}")
            
            # 返回文件下载路径
            download_path = output_path.replace(settings.DATA_DIR, "")
            if download_path.startswith('/'):
                download_path = download_path[1:]
            
            return {
                "success": True,
                "filename": filename,
                "download_path": f"/api/qrfile/download/{os.path.basename(output_path)}"
            }
        finally:
            # 清理临时目录
            shutil.rmtree(temp_dir, ignore_errors=True)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"恢复文件失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"恢复文件失败: {str(e)}")


@router.post("/scan-video")
async def scan_video(
    video_file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """
    从视频扫描二维码并恢复数据
    """
    try:
        # 检查文件类型
        filename = video_file.filename.lower()
        if not any(filename.endswith(ext) for ext in ['.mp4', '.avi', '.mov', '.mkv', '.wmv']):
            raise HTTPException(status_code=400, detail="不支持的视频格式，请上传MP4、AVI、MOV、MKV或WMV格式的视频")
        
        # 创建临时目录
        temp_dir = tempfile.mkdtemp(dir=QR_DIR)
        video_path = os.path.join(temp_dir, f"{uuid.uuid4()}{os.path.splitext(video_file.filename)[1]}")
        
        # 保存上传的视频文件
        with open(video_path, 'wb') as f:
            content = await video_file.read()
            f.write(content)
            
        # 检查文件大小
        file_size = os.path.getsize(video_path)
        logger.info(f"接收到视频文件: {video_file.filename}, 大小: {file_size/1024/1024:.2f}MB")
        
        # 检查视频文件是否可以打开
        try:
            cap = cv2.VideoCapture(video_path)
            if not cap.isOpened():
                cap.release()
                raise HTTPException(status_code=400, detail="无法打开视频文件，请检查文件是否损坏")
                
            # 获取视频基本信息
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            fps = cap.get(cv2.CAP_PROP_FPS)
            frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            cap.release()
            
            logger.info(f"视频信息: 分辨率={width}x{height}, FPS={fps:.2f}, 总帧数={frame_count}")
            
            if width < 320 or height < 240:
                logger.warning(f"视频分辨率较低: {width}x{height}，可能影响二维码识别")
        except Exception as e:
            logger.error(f"检查视频文件失败: {str(e)}")
            raise HTTPException(status_code=400, detail=f"视频文件无效: {str(e)}")
        
        try:
            # 创建视频扫描器并开始扫描
            # 由于视频处理可能耗时较长，使用后台线程处理
            def process_video():
                try:
                    logger.info(f"开始处理视频: {video_file.filename}")
                    scanner = VideoQRScanner(video_path, QR_DIR)
                    qr_data_list = scanner.start()
                    
                    if not qr_data_list:
                        logger.warning(f"视频中未找到二维码: {video_file.filename}")
                        return None, "视频中未找到二维码，请确保视频中包含清晰可见的二维码，或尝试使用图片扫描模式"
                    
                    logger.info(f"从视频中提取到 {len(qr_data_list)} 个二维码数据")
                    
                    # 合并二维码数据
                    processor = QRProcessor(QR_DIR)
                    try:
                        combined_data = processor.combine_data(qr_data_list)
                    except Exception as e:
                        logger.error(f"合并二维码数据失败: {str(e)}")
                        return None, f"合并二维码数据失败: {str(e)}"
                    
                    # 恢复数据
                    try:
                        output_path, filename = processor.restore(combined_data)
                    except Exception as e:
                        logger.error(f"恢复文件失败: {str(e)}")
                        return None, f"恢复文件失败: {str(e)}"
                    
                    logger.info(f"用户 {current_user.username} 成功从视频恢复文件: {filename}")
                    return output_path, filename
                except Exception as e:
                    logger.error(f"视频扫描处理失败: {str(e)}")
                    return None, str(e)
            
            # 创建后台任务处理视频
            loop = asyncio.get_event_loop()
            output_path, result_message = await loop.run_in_executor(None, process_video)
            
            # 清理临时文件
            try:
                os.unlink(video_path)
                os.rmdir(temp_dir)
            except:
                pass
            
            if not output_path:
                raise HTTPException(status_code=400, detail=f"处理视频失败: {result_message}")
            
            # 返回文件下载路径
            filename = os.path.basename(output_path)
            return {
                "success": True,
                "filename": filename,
                "download_path": f"/api/qrfile/download/{filename}"
            }
        finally:
            # 确保清理临时文件
            try:
                if os.path.exists(video_path):
                    os.unlink(video_path)
                if os.path.exists(temp_dir):
                    os.rmdir(temp_dir)
            except:
                pass
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"视频扫描失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"视频扫描失败: {str(e)}")


@router.get("/download/{filename}")
async def download_restored_file(
    filename: str,
    current_user: User = Depends(download_qrfile),
):
    """
    下载恢复的文件
    """
    try:
        file_path = os.path.join(QR_DIR, filename)
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="文件不存在")
            
        return FileResponse(
            file_path, 
            filename=filename,
            media_type="application/octet-stream"
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"下载文件失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"下载文件失败: {str(e)}")


@router.post("/get-excel-sheets")
async def get_excel_sheets(
    excel_file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """
    获取Excel文件的sheet列表
    """
    try:
        # 读取Excel文件内容
        excel_data = await excel_file.read()
        
        # 创建临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            temp_file.write(excel_data)
            excel_path = temp_file.name
        
        try:
            # 加载Excel
            from openpyxl import load_workbook
            wb = load_workbook(excel_path)
            sheet_names = wb.sheetnames
            
            # 清理临时文件
            os.unlink(excel_path)
            
            return {"success": True, "sheets": sheet_names}
        except Exception as e:
            # 清理临时文件
            try:
                os.unlink(excel_path)
            except:
                pass
            logger.error(f"解析Excel失败: {str(e)}")
            raise HTTPException(status_code=400, detail=f"无效的Excel文件: {str(e)}")
    except Exception as e:
        logger.error(f"获取Excel sheets失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"获取Excel sheets失败: {str(e)}")


@router.post("/restore-from-text")
async def restore_from_text(
    request: dict = Body(...),
    current_user: User = Depends(get_current_user),
):
    """
    从文本内容恢复文件
    文本内容是多个二维码的内容，使用英文分号(;)分隔
    """
    try:
        # 获取文本数据
        logger.info(f"收到文本恢复请求: {request}")
        text_data = request.get("text_data", "")
        if not text_data:
            logger.warning("请求中未提供文本数据")
            raise HTTPException(status_code=400, detail="未提供文本数据")
            
        # 创建临时目录
        temp_dir = tempfile.mkdtemp(dir=QR_DIR)
        
        try:
            # 分割文本内容
            chunks = text_data.split(';')
            chunks = [chunk.strip() for chunk in chunks if chunk.strip()]
            
            if not chunks:
                logger.warning("未提供有效的二维码文本数据")
                raise HTTPException(status_code=400, detail="未提供有效的二维码文本数据")
            
            logger.info(f"用户 {current_user.username} 提交了 {len(chunks)} 个文本块进行恢复")
            logger.debug(f"第一个块的前100个字符: {chunks[0][:100]}...")
            
            # 合并数据
            processor = QRProcessor(QR_DIR)
            try:
                combined_data = processor.combine_data(chunks)
                logger.info(f"成功合并数据，大小: {len(combined_data)} 字节")
            except Exception as e:
                logger.error(f"合并文本数据失败: {str(e)}")
                raise HTTPException(status_code=400, detail=f"合并文本数据失败: {str(e)}")
            
            # 恢复数据
            try:
                output_path, filename = processor.restore(combined_data)
                logger.info(f"成功恢复文件: {filename}, 路径: {output_path}")
            except Exception as e:
                logger.error(f"从文本恢复文件失败: {str(e)}")
                raise HTTPException(status_code=400, detail=f"从文本恢复文件失败: {str(e)}")
            
            logger.info(f"用户 {current_user.username} 成功从文本恢复文件: {filename}")
            
            # 返回文件下载路径
            return {
                "success": True,
                "filename": filename,
                "download_path": f"/api/qrfile/download/{os.path.basename(output_path)}"
            }
        finally:
            # 清理临时目录
            shutil.rmtree(temp_dir, ignore_errors=True)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"从文本恢复文件失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"从文本恢复文件失败: {str(e)}")


# 添加二维码文件管理API端点
@router.get("/files")
async def get_qr_files(
    skip: int = 0,
    limit: int = 10,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    获取用户的二维码文件列表
    """
    try:
        # 查询用户的二维码文件
        query = select(QRFile).where(
            QRFile.user_id == current_user.id,
            QRFile.is_deleted == False
        ).order_by(QRFile.created_at.desc()).offset(skip).limit(limit)
        
        result = await db.execute(query)
        qr_files = result.scalars().all()
        
        # 查询总数
        count_query = select(func.count()).select_from(QRFile).where(
            QRFile.user_id == current_user.id,
            QRFile.is_deleted == False
        )
        result = await db.execute(count_query)
        total_count = result.scalar()
        
        # 转换为响应格式
        files_data = []
        for qf in qr_files:
            # 检查二维码目录是否存在
            qr_dir = os.path.join(QR_DIR, qf.session_id)
            has_qr_codes = os.path.exists(qr_dir) and len(os.listdir(qr_dir)) > 0
            
            files_data.append({
                "id": qf.id,
                "session_id": qf.session_id,
                "mode": qf.mode,
                "original_filename": qf.original_filename,
                "original_filesize": qf.original_filesize,
                "chunk_count": qf.chunk_count,
                "has_qr_codes": has_qr_codes,
                "created_at": qf.created_at.isoformat(),
                "description": qf.description
            })
        
        return {
            "success": True,
            "files": files_data,
            "total": total_count,
            "skip": skip,
            "limit": limit
        }
    except Exception as e:
        logger.error(f"获取二维码文件列表失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"获取二维码文件列表失败: {str(e)}")


@router.delete("/files/{file_id}")
async def delete_qr_file(
    file_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    删除二维码文件
    """
    try:
        # 查询文件
        query = select(QRFile).where(
            QRFile.id == file_id,
            QRFile.user_id == current_user.id,
            QRFile.is_deleted == False
        )
        result = await db.execute(query)
        qr_file = result.scalars().first()
        
        if not qr_file:
            raise HTTPException(status_code=404, detail="文件不存在或已被删除")
        
        # 标记为已删除
        qr_file.is_deleted = True
        await db.commit()
        
        # 删除物理文件（可选，也可以保留一段时间后再清理）
        try:
            # 删除数据文件
            if os.path.exists(qr_file.filepath):
                os.unlink(qr_file.filepath)
                
            # 删除二维码目录
            qr_dir = os.path.join(QR_DIR, qr_file.session_id)
            if os.path.exists(qr_dir):
                shutil.rmtree(qr_dir)
        except Exception as e:
            logger.warning(f"删除物理文件失败: {str(e)}")
        
        return {
            "success": True,
            "message": "文件已成功删除"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除二维码文件失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"删除二维码文件失败: {str(e)}")


@router.post("/clean-old-files")
async def clean_old_files(
    days: int = Body(30),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    清理指定天数前的二维码文件（仅管理员可用）
    """
    try:
        # 检查权限
        if not current_user.is_superuser:
            raise HTTPException(status_code=403, detail="权限不足，仅管理员可执行此操作")
        
        # 计算截止日期
        cutoff_date = datetime.now() - timedelta(days=days)
        
        # 查询需要清理的文件
        query = select(QRFile).where(
            QRFile.created_at < cutoff_date,
            QRFile.is_deleted == False
        )
        result = await db.execute(query)
        old_files = result.scalars().all()
        
        # 标记为已删除并删除物理文件
        deleted_count = 0
        for qf in old_files:
            qf.is_deleted = True
            
            try:
                # 删除数据文件
                if os.path.exists(qf.filepath):
                    os.unlink(qf.filepath)
                    
                # 删除二维码目录
                qr_dir = os.path.join(QR_DIR, qf.session_id)
                if os.path.exists(qr_dir):
                    shutil.rmtree(qr_dir)
                    
                deleted_count += 1
            except Exception as e:
                logger.warning(f"删除物理文件失败: {str(e)}")
        
        await db.commit()
        
        return {
            "success": True,
            "deleted_count": deleted_count,
            "message": f"已清理 {deleted_count} 个超过 {days} 天的二维码文件"
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"清理旧文件失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"清理旧文件失败: {str(e)}")


@router.get("/file-stats")
async def get_file_stats(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    获取二维码文件统计信息
    """
    try:
        # 检查权限
        if not current_user.is_superuser:
            # 普通用户只能查看自己的统计信息
            # 查询用户的文件数量
            count_query = select(func.count()).select_from(QRFile).where(
                QRFile.user_id == current_user.id,
                QRFile.is_deleted == False
            )
            result = await db.execute(count_query)
            file_count = result.scalar()
            
            # 查询用户的文件总大小
            size_query = select(func.sum(QRFile.filesize)).select_from(QRFile).where(
                QRFile.user_id == current_user.id,
                QRFile.is_deleted == False
            )
            result = await db.execute(size_query)
            total_size = result.scalar() or 0
            
            return {
                "success": True,
                "file_count": file_count,
                "total_size": total_size,
                "total_size_human": format_size(total_size)
            }
        else:
            # 管理员可以查看所有统计信息
            # 查询总文件数
            count_query = select(func.count()).select_from(QRFile).where(
                QRFile.is_deleted == False
            )
            result = await db.execute(count_query)
            total_file_count = result.scalar()
            
            # 查询总文件大小
            size_query = select(func.sum(QRFile.filesize)).select_from(QRFile).where(
                QRFile.is_deleted == False
            )
            result = await db.execute(size_query)
            total_size = result.scalar() or 0
            
            # 查询用户数量
            user_query = select(func.count(QRFile.user_id.distinct())).select_from(QRFile).where(
                QRFile.is_deleted == False
            )
            result = await db.execute(user_query)
            user_count = result.scalar()
            
            # 查询已删除的文件数量
            deleted_query = select(func.count()).select_from(QRFile).where(
                QRFile.is_deleted == True
            )
            result = await db.execute(deleted_query)
            deleted_count = result.scalar()
            
            return {
                "success": True,
                "file_count": total_file_count,
                "total_size": total_size,
                "total_size_human": format_size(total_size),
                "user_count": user_count,
                "deleted_count": deleted_count
            }
    except Exception as e:
        logger.error(f"获取文件统计信息失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"获取文件统计信息失败: {str(e)}")


# 辅助函数：格式化文件大小
def format_size(size_bytes):
    """将字节大小格式化为人类可读的形式"""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.2f} KB"
    elif size_bytes < 1024 * 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.2f} MB"
    else:
        return f"{size_bytes / (1024 * 1024 * 1024):.2f} GB" 
    

# 根据session_id获取二维码文件列表
@router.get("/files/{session_id}")
async def get_qr_files_by_session_id(
    session_id: str,
    current_user: User = Depends(get_current_user),
):
    """
    根据session_id获取二维码文件列表
    """
    try:
        # 检查权限
        if not current_user.is_superuser:
            raise HTTPException(status_code=403, detail="权限不足，仅管理员可执行此操作")
        
        # 查询二维码文件
        qr_dir = os.path.join(QR_DIR, session_id)
        if not os.path.exists(qr_dir):
            raise HTTPException(status_code=404, detail="二维码文件不存在")

        # 获取文件列表
        files = os.listdir(qr_dir)
        return {
            "success": True,
            "files": files
        }
    except Exception as e:
        logger.error(f"根据session_id获取二维码文件列表失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"根据session_id获取二维码文件列表失败: {str(e)}")

